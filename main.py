import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


APP_NAME = "Acquisition Packet Generator"
APP_VERSION = "0.1"

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
SAVED_PACKETS_DIR = BASE_DIR / "saved_packets"
SETTINGS_PATH = BASE_DIR / "settings.json"
FPR_TRACKING_PATH = BASE_DIR / "fpr_tracking.xlsx"


DEFAULT_SETTINGS = {
    "department_name": "Example Police Department",
    "unit_name": "Digital Evidence Unit",
    "default_technician": "Matthew McBride",
    "common_technicians": [
        "Matthew McBride"
    ],
    "common_investigators": [
        "Officer Example",
        "Detective Example"
    ],
    "common_tools": [
        {
            "name": "Cellebrite UFED",
            "version": ""
        },
        {
            "name": "Cellebrite Physical Analyzer",
            "version": ""
        },
        {
            "name": "Magnet AXIOM",
            "version": ""
        },
        {
            "name": "FTK Imager",
            "version": ""
        },
        {
            "name": "GrayKey",
            "version": ""
        }
    ],
    "common_evidence_locations": [
        "Evidence Locker",
        "Property Room",
        "Digital Evidence Storage"
    ],
    "default_scope_statement": (
        "This documentation records the technical acquisition, extraction, imaging, "
        "and/or processing steps performed on the listed device or media. The technician "
        "did not conduct a content examination, investigative review, interpretation of user "
        "activity, or evidentiary analysis unless specifically stated. Any review, "
        "interpretation, or investigative conclusions are the responsibility of the assigned "
        "officer, investigator, or case agent."
    )
}


DEVICE_TYPES = [
    "CPU",
    "ETech",
    "Loose Drive",
    "Mobile Phone",
    "Tablet",
    "USB Drive",
    "SD Card",
    "DVR/NVR Storage",
    "Other"
]


PROCESSING_TYPES = [
    "Mobile Device Extraction",
    "Drive / Media Imaging",
    "Reader Report Generation",
    "Case File Generation",
    "Extraction Export",
    "Other"
]


PROCESSING_STATUSES = [
    "Completed",
    "Completed with Limitations",
    "Partial Extraction / Image",
    "Unable to Process",
    "Cancelled by Requestor",
    "Pending"
]


OUTPUT_TYPES = [
    "Reader Report",
    "Case File",
    "Forensic Image",
    "Extraction Export",
    "Other"
]


STORAGE_UNITS = [
    "GB",
    "TB",
    "MB",
    "Unknown"
]


def ensure_directories():
    OUTPUT_DIR.mkdir(exist_ok=True)
    SAVED_PACKETS_DIR.mkdir(exist_ok=True)


def load_or_create_settings():
    if not SETTINGS_PATH.exists():
        SETTINGS_PATH.write_text(
            json.dumps(DEFAULT_SETTINGS, indent=4),
            encoding="utf-8"
        )
        print(f"Created default settings file: {SETTINGS_PATH}")

    with SETTINGS_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def save_settings(settings):
    SETTINGS_PATH.write_text(
        json.dumps(settings, indent=4),
        encoding="utf-8"
    )


def ask_text(question, default=None, required=False):
    while True:
        if default:
            answer = input(f"{question} [{default}]: ").strip()
            if not answer:
                answer = default
        else:
            answer = input(f"{question}: ").strip()

        if required and not answer:
            print("This field is required.")
            continue

        return answer


def ask_yes_no(question, default="N"):
    default = default.upper()
    prompt = f"{question} (Y/N) [{default}]: "

    while True:
        answer = input(prompt).strip().upper()

        if not answer:
            answer = default

        if answer in ["Y", "YES"]:
            return True

        if answer in ["N", "NO"]:
            return False

        print("Please enter Y or N.")


def ask_choice(question, options, default=None):
    print(f"\n{question}")

    for index, option in enumerate(options, start=1):
        print(f"{index}. {option}")

    while True:
        if default:
            answer = input(f"Choose an option number [{default}]: ").strip()
            if not answer:
                answer = str(default)
        else:
            answer = input("Choose an option number: ").strip()

        if answer.isdigit():
            choice = int(answer)
            if 1 <= choice <= len(options):
                return options[choice - 1]

        print("Invalid choice. Try again.")


def ask_float(question, allow_blank=True):
    while True:
        answer = input(f"{question}: ").strip()

        if not answer and allow_blank:
            return None

        try:
            return float(answer)
        except ValueError:
            print("Enter a number, or leave blank if unknown.")


def ask_int(question, default=1):
    while True:
        answer = input(f"{question} [{default}]: ").strip()

        if not answer:
            return default

        if answer.isdigit() and int(answer) > 0:
            return int(answer)

        print("Enter a whole number greater than zero.")


def safe_filename(value):
    unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    safe_value = value.strip()

    for char in unsafe_chars:
        safe_value = safe_value.replace(char, "-")

    return safe_value.replace(" ", "_")


def convert_to_gb(size, unit):
    if size is None or unit == "Unknown":
        return None

    if unit == "GB":
        return size

    if unit == "TB":
        return size * 1024

    if unit == "MB":
        return size / 1024

    return None


def format_storage_gb(value):
    if value is None:
        return "Unknown"

    if value >= 1024:
        return f"{value:,.2f} GB ({value / 1024:,.2f} TB)"

    return f"{value:,.2f} GB"


def count_devices_by_type(devices):
    counts = {}

    for device in devices:
        device_type = device["device_type"]
        quantity = device["quantity"]
        counts[device_type] = counts.get(device_type, 0) + quantity

    return counts


def calculate_total_storage_gb(devices):
    total = 0
    has_known_storage = False

    for device in devices:
        storage_total_gb = device.get("storage_total_gb")

        if storage_total_gb is not None:
            total += storage_total_gb
            has_known_storage = True

    if not has_known_storage:
        return None

    return total


def collect_tools(settings):
    selected_tools = []
    common_tools = settings.get("common_tools", [])

    print("\nTools Used")
    print("Select tools used. Enter numbers separated by commas.")
    print("Example: 1,3,4")
    print("Leave blank if none of the listed tools apply.")

    for index, tool in enumerate(common_tools, start=1):
        name = tool.get("name", "Unnamed Tool")
        version = tool.get("version", "")
        version_display = f" v{version}" if version else ""
        print(f"{index}. {name}{version_display}")

    answer = input("Selected tools: ").strip()

    if answer:
        selections = [item.strip() for item in answer.split(",")]

        for selection in selections:
            if selection.isdigit():
                index = int(selection)
                if 1 <= index <= len(common_tools):
                    tool = common_tools[index - 1]
                    name = tool.get("name", "Unnamed Tool")
                    default_version = tool.get("version", "")
                    version = ask_text(f"Version for {name}", default=default_version)
                    selected_tools.append({
                        "name": name,
                        "version": version
                    })

    while ask_yes_no("Add another custom tool?", default="N"):
        name = ask_text("Tool name", required=True)
        version = ask_text("Tool version")
        selected_tools.append({
            "name": name,
            "version": version
        })

    return selected_tools


def collect_devices():
    devices = []

    print("\nDevices / Media")
    print("Enter each device category or evidence group.")

    while True:
        device_type = ask_choice("Device type", DEVICE_TYPES)

        if device_type == "Other":
            device_type = ask_text("Enter custom device type", required=True)

        quantity = ask_int("Quantity", default=1)

        description = ask_text("Description")
        make = ask_text("Make")
        model = ask_text("Model")
        serial = ask_text("Serial / identifier")
        capacity_size = ask_float("Storage size per device, leave blank if unknown")
        capacity_unit = "Unknown"

        if capacity_size is not None:
            capacity_unit = ask_choice("Storage unit", STORAGE_UNITS, default=1)

        storage_each_gb = convert_to_gb(capacity_size, capacity_unit)
        storage_total_gb = None

        if storage_each_gb is not None:
            storage_total_gb = storage_each_gb * quantity

        devices.append({
            "device_type": device_type,
            "quantity": quantity,
            "description": description,
            "make": make,
            "model": model,
            "serial": serial,
            "capacity_size": capacity_size,
            "capacity_unit": capacity_unit,
            "storage_each_gb": storage_each_gb,
            "storage_total_gb": storage_total_gb
        })

        if not ask_yes_no("Add another device/media item?", default="N"):
            break

    return devices


def collect_packet(settings):
    print("\nNew Acquisition Packet")
    print("=" * 50)

    packet = {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "department": {
            "department_name": settings.get("department_name", ""),
            "unit_name": settings.get("unit_name", "")
        },
        "general_info": {},
        "intake_info": {},
        "subject": {},
        "devices": [],
        "tools_used": [],
        "processing": {},
        "output": {},
        "scope_statement": settings.get("default_scope_statement", "")
    }

    print("\nGeneral Information")
    packet["general_info"]["case_number"] = ask_text("Case number", required=True)
    packet["general_info"]["agency_case_number"] = ask_text("Agency case number, if different")
    packet["general_info"]["offense_or_incident"] = ask_text("Offense / incident type")

    packet["subject"]["last_name"] = ask_text("Subject last name")
    packet["subject"]["first_name"] = ask_text("Subject first name")

    common_investigators = settings.get("common_investigators", [])
    if common_investigators and ask_yes_no("Select requesting investigator from defaults?", default="Y"):
        investigator = ask_choice("Requesting officer / investigator", common_investigators)
    else:
        investigator = ask_text("Requesting officer / investigator", required=True)

    packet["general_info"]["requesting_investigator"] = investigator

    common_technicians = settings.get("common_technicians", [])
    default_technician = settings.get("default_technician", "")

    if common_technicians and ask_yes_no("Select technician from defaults?", default="Y"):
        technician = ask_choice("Technician", common_technicians)
    else:
        technician = ask_text("Technician", default=default_technician, required=True)

    packet["general_info"]["technician"] = technician

    packet["general_info"]["date_received"] = ask_text("Date received")
    packet["general_info"]["time_received"] = ask_text("Time received")
    packet["general_info"]["date_processed"] = ask_text("Date processed")
    packet["general_info"]["time_processed"] = ask_text("Time processed")

    print("\nIntake Information")
    packet["intake_info"]["dropoff_person"] = ask_text("Drop-off person")
    packet["intake_info"]["received_from"] = ask_text("Received from")
    packet["intake_info"]["evidence_item_number"] = ask_text("Evidence item number")
    packet["intake_info"]["checked_out_from_evidence"] = ask_yes_no("Was item checked out from evidence?", default="Y")
    packet["intake_info"]["checked_out_datetime"] = ask_text("Checked out date/time")
    packet["intake_info"]["returned_to_evidence"] = ask_yes_no("Was item returned to evidence?", default="Y")
    packet["intake_info"]["returned_datetime"] = ask_text("Returned date/time")

    locations = settings.get("common_evidence_locations", [])
    if locations and ask_yes_no("Select evidence location from defaults?", default="Y"):
        evidence_location = ask_choice("Evidence location", locations)
    else:
        evidence_location = ask_text("Evidence location / locker")

    packet["intake_info"]["evidence_location"] = evidence_location

    packet["devices"] = collect_devices()
    packet["tools_used"] = collect_tools(settings)

    print("\nProcessing Information")
    packet["processing"]["processing_type"] = ask_choice("Processing type", PROCESSING_TYPES)
    packet["processing"]["processing_status"] = ask_choice("Processing status", PROCESSING_STATUSES)
    packet["processing"]["processing_notes"] = ask_text("Processing notes")

    print("\nOutput Information")
    packet["output"]["output_type"] = ask_choice("Output type", OUTPUT_TYPES)
    packet["output"]["output_filename"] = ask_text("Output filename / identifier")
    packet["output"]["output_location"] = ask_text("Output location")
    packet["output"]["reader_report_generated"] = ask_yes_no("Reader report generated?", default="Y")
    packet["output"]["case_file_generated"] = ask_yes_no("Case file generated?", default="N")

    print("\nTechnician Notes")
    packet["technician_notes"] = ask_text("Technician notes")

    packet["summary"] = {
        "device_counts": count_devices_by_type(packet["devices"]),
        "total_devices": sum(device["quantity"] for device in packet["devices"]),
        "total_storage_gb": calculate_total_storage_gb(packet["devices"])
    }

    return packet


def build_txt_report(packet):
    general = packet["general_info"]
    intake = packet["intake_info"]
    subject = packet["subject"]
    processing = packet["processing"]
    output = packet["output"]
    summary = packet["summary"]

    department_name = packet["department"].get("department_name", "")
    unit_name = packet["department"].get("unit_name", "")

    lines = []

    lines.append("DIGITAL EVIDENCE ACQUISITION DOCUMENTATION")
    lines.append("=" * 55)
    lines.append(department_name)
    lines.append(unit_name)
    lines.append("")

    lines.append("Administrative Information")
    lines.append("-" * 30)
    lines.append(f"Case Number: {general.get('case_number', '')}")
    lines.append(f"Agency Case Number: {general.get('agency_case_number', '')}")
    lines.append(f"Offense / Incident: {general.get('offense_or_incident', '')}")
    lines.append(f"Subject: {subject.get('last_name', '')}, {subject.get('first_name', '')}")
    lines.append(f"Requesting Officer / Investigator: {general.get('requesting_investigator', '')}")
    lines.append(f"Technician: {general.get('technician', '')}")
    lines.append(f"Date Received: {general.get('date_received', '')}")
    lines.append(f"Time Received: {general.get('time_received', '')}")
    lines.append(f"Date Processed: {general.get('date_processed', '')}")
    lines.append(f"Time Processed: {general.get('time_processed', '')}")
    lines.append("")

    lines.append("Intake Information")
    lines.append("-" * 30)
    lines.append(f"Drop-off Person: {intake.get('dropoff_person', '')}")
    lines.append(f"Received From: {intake.get('received_from', '')}")
    lines.append(f"Evidence Item Number: {intake.get('evidence_item_number', '')}")
    lines.append(f"Checked Out From Evidence: {intake.get('checked_out_from_evidence', '')}")
    lines.append(f"Checked Out Date/Time: {intake.get('checked_out_datetime', '')}")
    lines.append(f"Returned To Evidence: {intake.get('returned_to_evidence', '')}")
    lines.append(f"Returned Date/Time: {intake.get('returned_datetime', '')}")
    lines.append(f"Evidence Location / Locker: {intake.get('evidence_location', '')}")
    lines.append("")

    lines.append("Device / Media Summary")
    lines.append("-" * 30)
    lines.append(f"Total Devices / Media Count: {summary.get('total_devices', 0)}")
    lines.append(f"Total Known Storage: {format_storage_gb(summary.get('total_storage_gb'))}")
    lines.append("")

    for device_type, count in summary.get("device_counts", {}).items():
        lines.append(f"{device_type}: {count}")

    lines.append("")
    lines.append("Device / Media Detail")
    lines.append("-" * 30)

    for index, device in enumerate(packet["devices"], start=1):
        lines.append(f"Device Entry {index}")
        lines.append(f"  Type: {device.get('device_type', '')}")
        lines.append(f"  Quantity: {device.get('quantity', '')}")
        lines.append(f"  Description: {device.get('description', '')}")
        lines.append(f"  Make: {device.get('make', '')}")
        lines.append(f"  Model: {device.get('model', '')}")
        lines.append(f"  Serial / Identifier: {device.get('serial', '')}")

        size = device.get("capacity_size")
        unit = device.get("capacity_unit")

        if size is None:
            lines.append("  Storage Per Device: Unknown")
        else:
            lines.append(f"  Storage Per Device: {size} {unit}")

        lines.append(f"  Total Known Storage: {format_storage_gb(device.get('storage_total_gb'))}")
        lines.append("")

    lines.append("Tools Used")
    lines.append("-" * 30)

    if packet["tools_used"]:
        for tool in packet["tools_used"]:
            name = tool.get("name", "")
            version = tool.get("version", "")
            lines.append(f"{name} | Version: {version}")
    else:
        lines.append("No tools entered.")

    lines.append("")

    lines.append("Processing Information")
    lines.append("-" * 30)
    lines.append(f"Processing Type: {processing.get('processing_type', '')}")
    lines.append(f"Processing Status: {processing.get('processing_status', '')}")
    lines.append(f"Processing Notes: {processing.get('processing_notes', '')}")
    lines.append("")

    lines.append("Generated Output")
    lines.append("-" * 30)
    lines.append(f"Output Type: {output.get('output_type', '')}")
    lines.append(f"Output Filename / Identifier: {output.get('output_filename', '')}")
    lines.append(f"Output Location: {output.get('output_location', '')}")
    lines.append(f"Reader Report Generated: {output.get('reader_report_generated', '')}")
    lines.append(f"Case File Generated: {output.get('case_file_generated', '')}")
    lines.append("")

    lines.append("Technician Notes")
    lines.append("-" * 30)
    lines.append(packet.get("technician_notes", ""))
    lines.append("")

    lines.append("Limitations and Scope")
    lines.append("-" * 30)
    lines.append(packet.get("scope_statement", ""))
    lines.append("")

    lines.append("Technician Statement")
    lines.append("-" * 30)
    lines.append(
        "I documented the above process based on the actions performed, tool output, "
        "and information available at the time of processing."
    )
    lines.append("")

    lines.append("Report Generated")
    lines.append("-" * 30)
    lines.append(f"Generated On: {packet.get('created_at', '')}")
    lines.append(f"Generated By: {APP_NAME} v{APP_VERSION}")

    return "\n".join(lines)


def save_packet_outputs(packet):
    case_number = packet["general_info"]["case_number"]
    safe_case = safe_filename(case_number)

    txt_path = OUTPUT_DIR / f"{safe_case}_acquisition_packet.txt"
    json_path = SAVED_PACKETS_DIR / f"{safe_case}_acquisition_packet.json"

    report_text = build_txt_report(packet)

    txt_path.write_text(report_text, encoding="utf-8")

    json_path.write_text(
        json.dumps(packet, indent=4),
        encoding="utf-8"
    )

    return txt_path, json_path


def style_header_row(sheet):
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def get_or_create_workbook():
    if FPR_TRACKING_PATH.exists():
        workbook = load_workbook(FPR_TRACKING_PATH)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    if "Case Summary" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Case Summary")
        sheet.append([
            "Case Number",
            "Agency Case Number",
            "Subject Last Name",
            "Subject First Name",
            "Offense / Incident",
            "Requesting Investigator",
            "Technician",
            "Date Processed",
            "CPU Count",
            "ETech Count",
            "Loose Drive Count",
            "Mobile Phone Count",
            "Tablet Count",
            "USB Drive Count",
            "SD Card Count",
            "DVR/NVR Storage Count",
            "Other Count",
            "Total Devices",
            "Total Storage GB",
            "Processing Type",
            "Processing Status",
            "Output Type",
            "Report Generated"
        ])
        style_header_row(sheet)

    if "Device Detail" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Device Detail")
        sheet.append([
            "Case Number",
            "Subject Last Name",
            "Subject First Name",
            "Device Type",
            "Quantity",
            "Description",
            "Make",
            "Model",
            "Serial / Identifier",
            "Storage Size Per Device",
            "Storage Unit",
            "Storage GB Per Device",
            "Storage GB Total",
            "Date Processed",
            "Technician"
        ])
        style_header_row(sheet)

    return workbook


def append_to_fpr_tracking(packet):
    workbook = get_or_create_workbook()

    general = packet["general_info"]
    subject = packet["subject"]
    processing = packet["processing"]
    output = packet["output"]
    summary = packet["summary"]
    counts = summary.get("device_counts", {})

    case_summary = workbook["Case Summary"]

    other_count = 0
    known_summary_types = [
        "CPU",
        "ETech",
        "Loose Drive",
        "Mobile Phone",
        "Tablet",
        "USB Drive",
        "SD Card",
        "DVR/NVR Storage"
    ]

    for device_type, count in counts.items():
        if device_type not in known_summary_types:
            other_count += count

    case_summary.append([
        general.get("case_number", ""),
        general.get("agency_case_number", ""),
        subject.get("last_name", ""),
        subject.get("first_name", ""),
        general.get("offense_or_incident", ""),
        general.get("requesting_investigator", ""),
        general.get("technician", ""),
        general.get("date_processed", ""),
        counts.get("CPU", 0),
        counts.get("ETech", 0),
        counts.get("Loose Drive", 0),
        counts.get("Mobile Phone", 0),
        counts.get("Tablet", 0),
        counts.get("USB Drive", 0),
        counts.get("SD Card", 0),
        counts.get("DVR/NVR Storage", 0),
        other_count,
        summary.get("total_devices", 0),
        summary.get("total_storage_gb"),
        processing.get("processing_type", ""),
        processing.get("processing_status", ""),
        output.get("output_type", ""),
        packet.get("created_at", "")
    ])

    device_detail = workbook["Device Detail"]

    for device in packet["devices"]:
        device_detail.append([
            general.get("case_number", ""),
            subject.get("last_name", ""),
            subject.get("first_name", ""),
            device.get("device_type", ""),
            device.get("quantity", ""),
            device.get("description", ""),
            device.get("make", ""),
            device.get("model", ""),
            device.get("serial", ""),
            device.get("capacity_size"),
            device.get("capacity_unit", ""),
            device.get("storage_each_gb"),
            device.get("storage_total_gb"),
            general.get("date_processed", ""),
            general.get("technician", "")
        ])

    for sheet in workbook.worksheets:
        autofit_columns(sheet)

    workbook.save(FPR_TRACKING_PATH)

    return FPR_TRACKING_PATH


def show_menu():
    print("\n" + "=" * 60)
    print(f"{APP_NAME} v{APP_VERSION}")
    print("=" * 60)
    print("1. New acquisition packet")
    print("2. Show settings file location")
    print("3. Reset settings to default")
    print("4. Exit")


def reset_settings():
    save_settings(DEFAULT_SETTINGS)
    print(f"Settings reset to default: {SETTINGS_PATH}")


def main():
    ensure_directories()
    settings = load_or_create_settings()

    while True:
        show_menu()
        choice = input("Choose an option: ").strip()

        if choice == "1":
            packet = collect_packet(settings)

            print("\nReview Summary")
            print("-" * 30)
            print(f"Case Number: {packet['general_info']['case_number']}")
            print(f"Subject: {packet['subject']['last_name']}, {packet['subject']['first_name']}")
            print(f"Total Devices: {packet['summary']['total_devices']}")
            print(f"Total Known Storage: {format_storage_gb(packet['summary']['total_storage_gb'])}")
            print(f"Tools Used: {len(packet['tools_used'])}")

            if ask_yes_no("Save packet and update FPR tracking workbook?", default="Y"):
                txt_path, json_path = save_packet_outputs(packet)
                xlsx_path = append_to_fpr_tracking(packet)

                print("\nDone.")
                print(f"TXT report saved to: {txt_path}")
                print(f"JSON packet saved to: {json_path}")
                print(f"FPR tracking workbook updated: {xlsx_path}")
            else:
                print("Packet discarded. Nothing was saved.")

        elif choice == "2":
            print(f"\nSettings file: {SETTINGS_PATH}")
            print("You can edit this JSON file in VS Code to update defaults.")

        elif choice == "3":
            if ask_yes_no("Reset settings to default?", default="N"):
                reset_settings()
                settings = load_or_create_settings()

        elif choice == "4":
            print("Exiting.")
            break

        else:
            print("Invalid menu choice.")


if __name__ == "__main__":
    main()