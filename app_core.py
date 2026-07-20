import json
import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from settings_service import (
    APP_NAME,
    APP_SUBTITLE,
    APP_VERSION,
    SUITE_NAME,
    PUBLISHER_NAME,
    PRODUCT_DOMAIN,
    TOOL_FOLDER_NAME,
    BASE_DIR,
    SETTINGS_PATH,
    DEFAULT_SETTINGS,
    get_output_paths,
    get_output_root,
    ensure_directories,
    load_or_create_settings,
    save_settings,
)


DEVICE_TYPES = [
    "CPU",
    "ETech",
    "Mobile Phone",
    "Loose Drive",
    "Tablet",
    "USB Drive",
    "SD Card",
    "DVR/NVR Storage",
    "Cloud",
    "Storage Media",
    "Other"
]


PROCESSING_TYPES = [
    "Mobile Device Extraction",
    "Drive / Media Imaging",
    "Reader Report Generation",
    "Case File Generation",
    "Extraction Export",
    "Other"
]


PROCESSING_STATUSES = [
    "Completed",
    "Completed with Limitations",
    "Partial Extraction / Image",
    "Unable to Process",
    "Cancelled by Requestor",
    "Pending"
]


OUTPUT_TYPES = [
    "Reader Report",
    "Case File",
    "Forensic Image",
    "Extraction Export",
    "Other"
]


STORAGE_UNITS = [
    "GB",
    "TB",
    "MB",
    "Unknown"
]


CASE_SUMMARY_HEADERS = [
    "Case Number",
    "Agency Case Number",
    "Agency Dropping Item Off",
    "Subject Last Name",
    "Subject First Name",
    "Offense / Incident",
    "Requesting Investigator",
    "Technician",
    "Date Processed",
    "Date Received",
    "CPU Count",
    "ETech Count",
    "Mobile Phone Count",
    "Loose Drive Count",
    "Tablet Count",
    "USB Drive Count",
    "SD Card Count",
    "DVR/NVR Storage Count",
    "Cloud Count",
    "Storage Media Count",
    "Other Count",
    "Total Devices",
    "Total Storage GB",
    "Total Media Examined",
    "Hard Drive Credits",
    "ETech Credits",
    "Media Credits",
    "Processing Type",
    "Processing Status",
    "Output Type",
    "Report Generated"
]


DEVICE_DETAIL_HEADERS = [
    "Case Number",
    "Subject Last Name",
    "Subject First Name",
    "Device Type",
    "Quantity",
    "Description",
    "Make",
    "Model",
    "Serial / Identifier",
    "Storage Size Per Device",
    "Storage Unit",
    "Storage GB Per Device",
    "Storage GB Total",
    "Volumes Examined",
    "Volume Scale",
    "Encrypted",
    "Decrypted",
    "Tools Used to Decrypt",
    "Password Locked",
    "Password Unlocked",
    "Services Used to Unlock",
    "Where Device Stored",
    "Condition Delivered",
    "Device Photo",
    "Copied Device Photo",
    "Date Processed",
    "Date Received",
    "Technician"
]


FPR_CASE_INFO_HEADERS = [
    "Case Number",
    "Agency Case Number",
    "Agency Dropping Item Off",
    "State / Local Case No.",
    "Subject Last Name",
    "Subject First Name",
    "Case Type",
    "Offense / Incident",
    "City of Offense",
    "State of Offense",
    "Country of Offense",
    "Exam Start Date",
    "Exam End Date",
    "Other Data Analyzed",
    "Case Summary",
    "Requesting Investigator",
    "Technician",
    "Date Processed",
    "Date Received",
    "Report Generated"
]


FPR_MEDIA_EXAMINED_HEADERS = [
    "Case Number",
    "Agency Case Number",
    "Agency Dropping Item Off",
    "Subject Last Name",
    "Subject First Name",
    "Device Type",
    "Quantity",
    "Volumes Examined",
    "Volume Scale",
    "Encrypted",
    "Decrypted",
    "Tools Used to Decrypt",
    "Password Locked",
    "Password Unlocked",
    "Services Used to Unlock",
    "Technician",
    "Date Processed",
    "Date Received",
    "Report Generated"
]


def safe_filename(value):
    """
    Convert a case number, title, timestamp, or other text value into a safer filename.
    """
    unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    safe_value = str(value).strip()

    for char in unsafe_chars:
        safe_value = safe_value.replace(char, "-")

    return safe_value.replace(" ", "_")


def convert_to_gb(size, unit):
    """
    Convert a storage size to GB.

    Returns None when size is unknown or cannot be converted.
    """
    if size is None or unit == "Unknown":
        return None

    if unit == "GB":
        return size

    if unit == "TB":
        return size * 1024

    if unit == "MB":
        return size / 1024

    return None


def format_storage_gb(value):
    """
    Format a GB value for report output.
    """
    if value is None:
        return "Unknown"

    if value >= 1024:
        return f"{value:,.2f} GB ({value / 1024:,.2f} TB)"

    return f"{value:,.2f} GB"


def yes_no(value):
    """
    Convert booleans to Yes/No for reports and spreadsheets.
    """
    return "Yes" if value else "No"




def is_supported_image_path(path_text):
    suffix = Path(str(path_text or "")).suffix.lower()
    return suffix in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff"}


def normalize_device_photos(device):
    """Return a normalized list of device photo dictionaries.

    v0.9.2 stored a single optional path in device_photo_path. v0.9.3 stores
    multiple labeled images in device_photos while preserving the legacy fields
    for older JSON compatibility.
    """
    photos = device.get("device_photos", [])
    if not isinstance(photos, list):
        photos = []

    normalized = []
    for index, photo in enumerate(photos, start=1):
        if isinstance(photo, dict):
            path_text = str(photo.get("path", "")).strip()
            label = str(photo.get("label", "")).strip()
        else:
            path_text = str(photo).strip()
            label = ""

        if path_text:
            normalized.append({
                "label": label or f"Photo {index}",
                "path": path_text,
                "copied_path": str(photo.get("copied_path", "")).strip() if isinstance(photo, dict) else "",
                "copy_note": str(photo.get("copy_note", "")).strip() if isinstance(photo, dict) else ""
            })

    legacy_path = str(device.get("device_photo_path", "")).strip()
    if legacy_path and not any(str(photo.get("path", "")).strip() == legacy_path for photo in normalized):
        normalized.append({
            "label": "Photo 1",
            "path": legacy_path,
            "copied_path": str(device.get("device_photo_copied_path", "")).strip(),
            "copy_note": str(device.get("device_photo_copy_note", "")).strip()
        })

    return normalized


def copy_device_photos(packet, settings=None):
    """Copy labeled device photos into the Acquire packet attachments folder.

    Missing or unsupported files are non-blocking because acquisition
    documentation should still export even if optional photos are unavailable.
    """
    general = packet.get("general_info", {})
    case_number = general.get("case_number", "UNKNOWN_CASE")
    paths = get_output_paths(settings, case_number=case_number)
    photo_dir = paths["device_photos_dir"]
    photo_dir.mkdir(parents=True, exist_ok=True)

    for device_index, device in enumerate(packet.get("devices", []), start=1):
        photos = normalize_device_photos(device)
        copied_photos = []

        for photo_index, photo in enumerate(photos, start=1):
            label = str(photo.get("label", "")).strip() or f"Photo {photo_index}"
            original = str(photo.get("path", "")).strip()

            record = {
                "label": label,
                "path": original,
                "copied_path": "",
                "copy_note": ""
            }

            if not original:
                copied_photos.append(record)
                continue

            source = Path(original)
            record["path"] = str(source)

            if not source.exists():
                record["copy_note"] = "Original photo path was not found during export."
                copied_photos.append(record)
                continue

            if not is_supported_image_path(source):
                record["copy_note"] = "Photo was not copied because the file type is not a supported image type."
                copied_photos.append(record)
                continue

            serial = safe_filename(device.get("serial", "") or f"device_{device_index:03d}")
            device_type = safe_filename(device.get("device_type", "device") or "device")
            label_name = safe_filename(label or f"photo_{photo_index:02d}")
            destination = photo_dir / f"device_{device_index:03d}_{device_type}_{serial}_{photo_index:02d}_{label_name}{source.suffix.lower()}"

            try:
                shutil.copy2(source, destination)
                record["copied_path"] = str(destination)
            except Exception as error:
                record["copy_note"] = f"Unable to copy device photo: {error}"

            copied_photos.append(record)

        device["device_photos"] = copied_photos

        # Preserve v0.9.2 fields for old templates/exports and existing user expectations.
        first_photo = copied_photos[0] if copied_photos else {}
        device["device_photo_path"] = first_photo.get("path", "")
        device["device_photo_copied_path"] = first_photo.get("copied_path", "")
        device["device_photo_copy_note"] = first_photo.get("copy_note", "")

    return packet


def count_devices_by_type(devices):
    """
    Count total device quantities by device type.
    """
    counts = {}

    for device in devices:
        device_type = device.get("device_type", "Other")
        quantity = device.get("quantity", 0)

        counts[device_type] = counts.get(device_type, 0) + quantity

    return counts


def calculate_total_storage_gb(devices):
    """
    Calculate total known storage across all device entries.

    Unknown storage entries are ignored.
    Returns None if no devices have known storage.
    """
    total = 0
    has_known_storage = False

    for device in devices:
        storage_total_gb = device.get("storage_total_gb")

        if storage_total_gb is not None:
            total += storage_total_gb
            has_known_storage = True

    if not has_known_storage:
        return None

    return total


def calculate_media_examined_summary(devices):
    """
    Derive FPR media examined values from the device/media table.

    Current working rules:
        Total Media Examined = total quantity of all device/media entries.
        Hard Drive Credits = CPU + Loose Drive + DVR/NVR Storage.
        ETech Credits = ETech + Mobile Phone.
        Media Credits = Tablet + USB Drive + SD Card + Cloud + Storage Media + Other.

    These rules are intentionally centralized here so the GUI does not ask
    the user to enter duplicate credit values manually.
    """
    counts = count_devices_by_type(devices)

    hard_drive_types = [
        "CPU",
        "Loose Drive",
        "DVR/NVR Storage"
    ]

    etech_types = [
        "ETech",
        "Mobile Phone"
    ]

    media_types = [
        "Tablet",
        "USB Drive",
        "SD Card",
        "Cloud",
        "Storage Media",
        "Other"
    ]

    total_media_examined = sum(device.get("quantity", 0) for device in devices)

    hard_drive_credits = sum(counts.get(device_type, 0) for device_type in hard_drive_types)
    etech_credits = sum(counts.get(device_type, 0) for device_type in etech_types)
    media_credits = sum(counts.get(device_type, 0) for device_type in media_types)

    known_credit_types = set(hard_drive_types + etech_types + media_types)

    for device_type, quantity in counts.items():
        if device_type not in known_credit_types:
            media_credits += quantity

    return {
        "total_media_examined": total_media_examined,
        "hard_drive_credits": hard_drive_credits,
        "etech_credits": etech_credits,
        "media_credits": media_credits
    }


def add_summary_to_packet(packet):
    """
    Add or refresh calculated summary sections of a packet.

    This makes the packet JSON, TXT output, and XLSX output consistent.
    The GUI should collect source data. Derived values belong here.
    """
    devices = packet.get("devices", [])

    packet["summary"] = {
        "device_counts": count_devices_by_type(devices),
        "total_devices": sum(device.get("quantity", 0) for device in devices),
        "total_storage_gb": calculate_total_storage_gb(devices)
    }

    packet["media_examined_summary"] = calculate_media_examined_summary(devices)

    return packet


def build_txt_report(packet):
    """
    Build the plain text acquisition packet report.
    """
    packet = add_summary_to_packet(packet)

    general = packet.get("general_info", {})
    intake = packet.get("intake_info", {})
    subject = packet.get("subject", {})
    processing = packet.get("processing", {})
    output = packet.get("output", {})
    report_info = packet.get("report_info", {})
    summary = packet.get("summary", {})
    media_summary = packet.get("media_examined_summary", {})

    department = packet.get("department", {})
    department_name = department.get("department_name", "")
    unit_name = department.get("unit_name", "")

    lines = []

    lines.append("BYTECASE ACQUIRE")
    lines.append("Digital Evidence Acquisition Documentation")
    lines.append("=" * 55)
    lines.append(f"Part of the {SUITE_NAME} toolset by {PUBLISHER_NAME}")
    lines.append(f"Product Domain: {PRODUCT_DOMAIN}")

    if department_name:
        lines.append(department_name)

    if unit_name:
        lines.append(unit_name)

    lines.append("")

    lines.append("Administrative Information")
    lines.append("-" * 30)
    lines.append(f"Case Number: {general.get('case_number', '')}")
    lines.append(f"Agency Case Number: {general.get('agency_case_number', '')}")
    lines.append(f"State / Local Case No.: {general.get('state_local_case_number', '')}")
    lines.append(f"Case Type: {general.get('case_type', '')}")
    lines.append(f"Offense / Incident: {general.get('offense_or_incident', '')}")
    lines.append(f"City of Offense: {general.get('city_of_offense', '')}")
    lines.append(f"State of Offense: {general.get('state_of_offense', '')}")
    lines.append(f"Country of Offense: {general.get('country_of_offense', '')}")
    lines.append(f"Subject: {subject.get('last_name', '')}, {subject.get('first_name', '')}")
    lines.append(f"Requesting Officer / Investigator: {general.get('requesting_investigator', '')}")
    lines.append(f"Technician: {general.get('technician', '')}")
    lines.append(f"Date Received: {general.get('date_received', '')}")
    lines.append(f"Time Received: {general.get('time_received', '')}")
    lines.append(f"Date Processed: {general.get('date_processed', '')}")
    lines.append(f"Time Processed: {general.get('time_processed', '')}")
    lines.append("")

    lines.append("Intake Information")
    lines.append("-" * 30)
    lines.append(f"Agency Dropping Item Off: {intake.get('agency_dropping_item_off', '')}")
    lines.append(f"Drop-off Person: {intake.get('dropoff_person', '')}")
    lines.append(f"Received From: {intake.get('received_from', '')}")
    lines.append(f"Evidence Item Number: {intake.get('evidence_item_number', '')}")
    lines.append(f"Checked Out From Evidence: {intake.get('checked_out_from_evidence', '')}")
    lines.append(f"Checked Out Date/Time: {intake.get('checked_out_datetime', '')}")
    lines.append(f"Returned To Evidence: {intake.get('returned_to_evidence', '')}")
    lines.append(f"Returned Date/Time: {intake.get('returned_datetime', '')}")
    lines.append(f"Evidence Location / Locker: {intake.get('evidence_location', '')}")
    lines.append("")

    lines.append("Examination Information")
    lines.append("-" * 30)
    lines.append(f"Exam Start Date: {processing.get('exam_start_date', '')}")
    lines.append(f"Exam End Date: {processing.get('exam_end_date', '')}")
    lines.append(f"Processing Type: {processing.get('processing_type', '')}")
    lines.append(f"Processing Status: {processing.get('processing_status', '')}")
    lines.append("Processing / Acquisition Narrative:")
    notes_text = processing.get("processing_notes", "")
    if notes_text:
        for note_line in str(notes_text).splitlines():
            lines.append(f"  {note_line}")
    else:
        lines.append("  ")
    lines.append("")

    lines.append("Device / Media Summary")
    lines.append("-" * 30)
    lines.append(f"Total Devices / Media Count: {summary.get('total_devices', 0)}")
    lines.append(f"Total Known Storage: {format_storage_gb(summary.get('total_storage_gb'))}")
    lines.append(f"Total Media Examined: {media_summary.get('total_media_examined', 0)}")
    lines.append(f"Hard Drive Credits: {media_summary.get('hard_drive_credits', 0)}")
    lines.append(f"ETech Credits: {media_summary.get('etech_credits', 0)}")
    lines.append(f"Media Credits: {media_summary.get('media_credits', 0)}")
    lines.append("")

    device_counts = summary.get("device_counts", {})

    if device_counts:
        for device_type, count in device_counts.items():
            lines.append(f"{device_type}: {count}")
    else:
        lines.append("No device/media entries recorded.")

    lines.append("")
    lines.append("Device / Media Detail")
    lines.append("-" * 30)

    devices = packet.get("devices", [])

    if devices:
        for index, device in enumerate(devices, start=1):
            lines.append(f"Device Entry {index}")
            lines.append(f"  Type: {device.get('device_type', '')}")
            lines.append(f"  Quantity: {device.get('quantity', '')}")
            lines.append(f"  Description: {device.get('description', '')}")
            lines.append(f"  Make: {device.get('make', '')}")
            lines.append(f"  Model: {device.get('model', '')}")
            lines.append(f"  Serial / Identifier: {device.get('serial', '')}")

            size = device.get("capacity_size")
            unit = device.get("capacity_unit")

            if size is None:
                lines.append("  Storage Per Device: Unknown")
            else:
                lines.append(f"  Storage Per Device: {size} {unit}")

            lines.append(f"  Total Known Storage: {format_storage_gb(device.get('storage_total_gb'))}")
            lines.append(f"  Volumes Examined: {device.get('volumes_examined', '')}")
            lines.append(f"  Volume Scale: {device.get('volume_scale', '')}")
            lines.append(f"  Encrypted: {yes_no(device.get('encrypted', False))}")
            lines.append(f"  Decrypted: {yes_no(device.get('decrypted', False))}")
            lines.append(f"  Tools Used to Decrypt: {device.get('tools_used_to_decrypt', '')}")
            lines.append(f"  Password Locked: {yes_no(device.get('password_locked', False))}")
            lines.append(f"  Password Unlocked: {yes_no(device.get('password_unlocked', False))}")
            lines.append(f"  Services Used to Unlock: {device.get('services_used_to_unlock', '')}")
            lines.append(f"  Where Device Stored: {device.get('device_storage_location', '')}")
            lines.append(f"  Condition Delivered: {device.get('condition_delivered', '')}")
            photos = normalize_device_photos(device)
            if photos:
                lines.append("  Device Photos:")
                for photo_index, photo in enumerate(photos, start=1):
                    lines.append(f"    {photo_index}. {photo.get('label', '')}")
                    lines.append(f"       Original Path: {photo.get('path', '')}")
                    lines.append(f"       Copied Path: {photo.get('copied_path', '')}")
                    if photo.get("copy_note"):
                        lines.append(f"       Note: {photo.get('copy_note', '')}")
            else:
                lines.append("  Device Photos: None recorded")
            lines.append("")
    else:
        lines.append("No device/media entries recorded.")
        lines.append("")

    lines.append("Tools Used")
    lines.append("-" * 30)

    tools_used = packet.get("tools_used", [])

    if tools_used:
        for tool in tools_used:
            name = tool.get("name", "")
            version = tool.get("version", "")
            lines.append(f"{name} | Version: {version}")
    else:
        lines.append("No tools entered.")

    lines.append("")

    lines.append("Generated Output")
    lines.append("-" * 30)
    lines.append(f"Output Type: {output.get('output_type', '')}")
    lines.append(f"Output Filename / Identifier: {output.get('output_filename', '')}")
    lines.append(f"Output Location: {output.get('output_location', '')}")
    lines.append(f"Reader Report Generated: {output.get('reader_report_generated', '')}")
    lines.append(f"Case File Generated: {output.get('case_file_generated', '')}")
    lines.append("")

    lines.append("Other Data Analyzed")
    lines.append("-" * 30)
    lines.append(report_info.get("other_data_analyzed", ""))
    lines.append("")

    lines.append("Case Summary")
    lines.append("-" * 30)
    lines.append(report_info.get("case_summary", ""))
    lines.append("")

    lines.append("Technician Notes")
    lines.append("-" * 30)
    lines.append(packet.get("technician_notes", ""))
    lines.append("")

    lines.append("Limitations and Scope")
    lines.append("-" * 30)
    lines.append(packet.get("scope_statement", ""))
    lines.append("")

    lines.append("Technician Statement")
    lines.append("-" * 30)
    lines.append(
        "I documented the above process based on the actions performed, tool output, "
        "and information available at the time of processing."
    )
    lines.append("")

    lines.append("Report Generated")
    lines.append("-" * 30)
    lines.append(f"Generated On: {packet.get('created_at', '')}")
    lines.append(f"Generated By: {APP_NAME} - {APP_SUBTITLE} v{APP_VERSION}")

    return "\n".join(lines)


def save_packet_outputs(packet, settings=None):
    """
    Save the acquisition packet as TXT and JSON.

    Filenames include case number, case title/offense value, and timestamp
    so repeated exports do not overwrite earlier packet files.

    Returns:
        tuple: (txt_path, json_path)
    """
    packet = add_summary_to_packet(packet)
    packet = copy_device_photos(packet, settings=settings)

    general = packet.get("general_info", {})
    case_number = general.get("case_number", "UNKNOWN_CASE")

    ensure_directories(settings, case_number=case_number)
    paths = get_output_paths(settings, case_number=case_number)
    report_info = packet.get("report_info", {})

    case_title = (
        general.get("offense_or_incident", "")
        or report_info.get("case_type", "")
        or general.get("case_type", "")
        or "Untitled_Case"
    )

    timestamp = packet.get("created_at", "")
    if not timestamp:
        timestamp = "unknown_time"

    safe_case = safe_filename(case_number)
    safe_title = safe_filename(case_title)
    safe_timestamp = safe_filename(timestamp)

    base_filename = f"{safe_case}_{safe_title}_{safe_timestamp}_acquisition_packet"

    txt_path = paths["reports_dir"] / f"{base_filename}.txt"
    json_path = paths["saved_packets_dir"] / f"{base_filename}.json"

    report_text = build_txt_report(packet)

    txt_path.write_text(report_text, encoding="utf-8")

    json_path.write_text(
        json.dumps(packet, indent=4),
        encoding="utf-8"
    )

    return txt_path, json_path


def style_header_row(sheet):
    """
    Apply basic styling to the first row of an XLSX sheet.
    """
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(sheet):
    """
    Set reasonable column widths based on cell contents.
    """
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


def ensure_sheet(workbook, sheet_name, headers):
    """
    Create a sheet if missing and ensure its header row matches the current schema.

    This lets existing tracking workbooks receive new columns without requiring
    the user to delete the workbook during development.
    """
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook[sheet_name]

    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index).value = header

    style_header_row(sheet)

    return sheet


def get_or_create_workbook(settings=None, case_number=None):
    """
    Open the existing case-specific FPR tracking workbook or create a new one.
    """
    paths = get_output_paths(settings, case_number=case_number)
    tracking_path = paths["tracking_workbook_path"]

    if tracking_path.exists():
        workbook = load_workbook(tracking_path)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    ensure_sheet(workbook, "Case Summary", CASE_SUMMARY_HEADERS)
    ensure_sheet(workbook, "Device Detail", DEVICE_DETAIL_HEADERS)
    ensure_sheet(workbook, "FPR Case Info", FPR_CASE_INFO_HEADERS)
    ensure_sheet(workbook, "FPR Media Examined", FPR_MEDIA_EXAMINED_HEADERS)

    return workbook


def append_to_fpr_tracking(packet, settings=None):
    """
    Append acquisition packet summary data to fpr_tracking.xlsx.

    The workbook contains:
    - Case Summary
    - Device Detail
    - FPR Case Info
    - FPR Media Examined

    Returns:
        Path: path to fpr_tracking.xlsx
    """
    packet = add_summary_to_packet(packet)
    packet = copy_device_photos(packet, settings=settings)

    general = packet.get("general_info", {})
    case_number = general.get("case_number", "UNKNOWN_CASE")

    ensure_directories(settings, case_number=case_number)
    paths = get_output_paths(settings, case_number=case_number)
    workbook = get_or_create_workbook(settings, case_number=case_number)
    subject = packet.get("subject", {})
    processing = packet.get("processing", {})
    output = packet.get("output", {})
    report_info = packet.get("report_info", {})
    media_examined_summary = packet.get("media_examined_summary", {})
    summary = packet.get("summary", {})
    counts = summary.get("device_counts", {})

    case_summary = workbook["Case Summary"]

    known_summary_types = [
        "CPU",
        "ETech",
        "Mobile Phone",
        "Loose Drive",
        "Tablet",
        "USB Drive",
        "SD Card",
        "DVR/NVR Storage",
        "Cloud",
        "Storage Media"
    ]

    other_count = 0

    for device_type, count in counts.items():
        if device_type not in known_summary_types:
            other_count += count

    case_summary.append([
        general.get("case_number", ""),
        general.get("agency_case_number", ""),
        packet.get("intake_info", {}).get("agency_dropping_item_off", ""),
        subject.get("last_name", ""),
        subject.get("first_name", ""),
        general.get("offense_or_incident", ""),
        general.get("requesting_investigator", ""),
        general.get("technician", ""),
        general.get("date_processed", ""),
        general.get("date_received", ""),
        counts.get("CPU", 0),
        counts.get("ETech", 0),
        counts.get("Mobile Phone", 0),
        counts.get("Loose Drive", 0),
        counts.get("Tablet", 0),
        counts.get("USB Drive", 0),
        counts.get("SD Card", 0),
        counts.get("DVR/NVR Storage", 0),
        counts.get("Cloud", 0),
        counts.get("Storage Media", 0),
        other_count,
        summary.get("total_devices", 0),
        summary.get("total_storage_gb"),
        media_examined_summary.get("total_media_examined", 0),
        media_examined_summary.get("hard_drive_credits", 0),
        media_examined_summary.get("etech_credits", 0),
        media_examined_summary.get("media_credits", 0),
        processing.get("processing_type", ""),
        processing.get("processing_status", ""),
        output.get("output_type", ""),
        packet.get("created_at", "")
    ])

    device_detail = workbook["Device Detail"]

    for device in packet.get("devices", []):
        device_detail.append([
            general.get("case_number", ""),
            subject.get("last_name", ""),
            subject.get("first_name", ""),
            device.get("device_type", ""),
            device.get("quantity", ""),
            device.get("description", ""),
            device.get("make", ""),
            device.get("model", ""),
            device.get("serial", ""),
            device.get("capacity_size"),
            device.get("capacity_unit", ""),
            device.get("storage_each_gb"),
            device.get("storage_total_gb"),
            device.get("volumes_examined", ""),
            device.get("volume_scale", ""),
            yes_no(device.get("encrypted", False)),
            yes_no(device.get("decrypted", False)),
            device.get("tools_used_to_decrypt", ""),
            yes_no(device.get("password_locked", False)),
            yes_no(device.get("password_unlocked", False)),
            device.get("services_used_to_unlock", ""),
            device.get("device_storage_location", ""),
            device.get("condition_delivered", ""),
            "; ".join(f"{photo.get('label', '')}: {photo.get('path', '')}" for photo in normalize_device_photos(device)),
            "; ".join(f"{photo.get('label', '')}: {photo.get('copied_path', '')}" for photo in normalize_device_photos(device)),
            general.get("date_processed", ""),
            general.get("technician", "")
        ])

    fpr_case_info = workbook["FPR Case Info"]

    fpr_case_info.append([
        general.get("case_number", ""),
        general.get("agency_case_number", ""),
        packet.get("intake_info", {}).get("agency_dropping_item_off", ""),
        report_info.get("state_local_case_number", general.get("state_local_case_number", "")),
        subject.get("last_name", ""),
        subject.get("first_name", ""),
        report_info.get("case_type", general.get("case_type", "")),
        general.get("offense_or_incident", ""),
        report_info.get("city_of_offense", general.get("city_of_offense", "")),
        report_info.get("state_of_offense", general.get("state_of_offense", "")),
        report_info.get("country_of_offense", general.get("country_of_offense", "")),
        report_info.get("exam_start_date", processing.get("exam_start_date", "")),
        report_info.get("exam_end_date", processing.get("exam_end_date", "")),
        report_info.get("other_data_analyzed", ""),
        report_info.get("case_summary", ""),
        general.get("requesting_investigator", ""),
        general.get("technician", ""),
        general.get("date_processed", ""),
        packet.get("created_at", "")
    ])

    fpr_media_examined = workbook["FPR Media Examined"]

    for device in packet.get("devices", []):
        fpr_media_examined.append([
            general.get("case_number", ""),
            general.get("agency_case_number", ""),
            subject.get("last_name", ""),
            subject.get("first_name", ""),
            device.get("device_type", ""),
            device.get("quantity", ""),
            device.get("volumes_examined", ""),
            device.get("volume_scale", ""),
            yes_no(device.get("encrypted", False)),
            yes_no(device.get("decrypted", False)),
            device.get("tools_used_to_decrypt", ""),
            yes_no(device.get("password_locked", False)),
            yes_no(device.get("password_unlocked", False)),
            device.get("services_used_to_unlock", ""),
            general.get("technician", ""),
            general.get("date_processed", ""),
            packet.get("created_at", "")
        ])

    for sheet in workbook.worksheets:
        autofit_columns(sheet)

    try:
        workbook.save(paths["tracking_workbook_path"])
    except PermissionError as error:
        raise PermissionError(
            "Unable to update the XLSX tracking workbook.\n\n"
            "The tracking workbook may already be open in Excel or locked by another program.\n\n"
            "Close fpr_tracking.xlsx and try generating the packet again."
        ) from error

    return paths["tracking_workbook_path"]

def parse_date_value(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def iter_saved_packet_json(settings=None):
    root = get_output_root(settings)
    pattern = f"*/{TOOL_FOLDER_NAME}/saved_packets/*.json"
    yield from root.glob(pattern)


def build_lab_statistics(settings=None, start_date="", end_date=""):
    start = parse_date_value(start_date)
    end = parse_date_value(end_date)

    rows = []
    totals = {
        "packets": 0,
        "devices": 0,
        "known_storage_gb": 0.0,
    }

    for json_path in iter_saved_packet_json(settings):
        try:
            packet = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        packet = add_summary_to_packet(packet)
        general = packet.get("general_info", {})
        subject = packet.get("subject", {})
        intake = packet.get("intake_info", {})
        processing = packet.get("processing", {})
        output = packet.get("output", {})
        summary = packet.get("summary", {})

        processed = parse_date_value(general.get("date_processed")) or parse_date_value(processing.get("exam_end_date")) or parse_date_value(general.get("date_received"))
        if start and processed and processed < start:
            continue
        if end and processed and processed > end:
            continue
        if start and not processed:
            continue
        if end and not processed:
            continue

        device_count = int(summary.get("total_devices", 0) or 0)
        storage = summary.get("total_storage_gb")
        storage_value = float(storage) if storage is not None else 0.0

        totals["packets"] += 1
        totals["devices"] += device_count
        totals["known_storage_gb"] += storage_value

        rows.append({
            "Date Processed": general.get("date_processed", ""),
            "Date Received": general.get("date_received", ""),
            "Case Number": general.get("case_number", ""),
            "Agency Case Number": general.get("agency_case_number", ""),
            "Agency Dropping Item Off": intake.get("agency_dropping_item_off", ""),
            "Subject Last Name": subject.get("last_name", ""),
            "Subject First Name": subject.get("first_name", ""),
            "Technician": general.get("technician", ""),
            "Processing Type": processing.get("processing_type", ""),
            "Processing Status": processing.get("processing_status", ""),
            "Output Type": output.get("output_type", ""),
            "Device Count": device_count,
            "Known Storage GB": storage if storage is not None else "",
            "Packet JSON": str(json_path),
        })

    return rows, totals


def export_lab_statistics(settings=None, start_date="", end_date=""):
    rows, totals = build_lab_statistics(settings=settings, start_date=start_date, end_date=end_date)
    ensure_directories(settings)
    paths = get_output_paths(settings)
    stats_dir = paths["lab_statistics_dir"]
    stats_dir.mkdir(parents=True, exist_ok=True)

    timestamp = safe_filename(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    suffix = "lab_statistics"
    if start_date or end_date:
        suffix = f"lab_statistics_{safe_filename(start_date or 'start')}_to_{safe_filename(end_date or 'end')}"

    csv_path = stats_dir / f"{timestamp}_{suffix}.csv"
    xlsx_path = stats_dir / f"{timestamp}_{suffix}.xlsx"

    headers = [
        "Date Processed", "Date Received", "Case Number", "Agency Case Number",
        "Agency Dropping Item Off", "Subject Last Name", "Subject First Name",
        "Technician", "Processing Type", "Processing Status", "Output Type",
        "Device Count", "Known Storage GB", "Packet JSON"
    ]

    with csv_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lab Statistics"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    totals_sheet = workbook.create_sheet("Summary")
    totals_sheet.append(["Metric", "Value"])
    totals_sheet.append(["Packets", totals.get("packets", 0)])
    totals_sheet.append(["Devices", totals.get("devices", 0)])
    totals_sheet.append(["Known Storage GB", totals.get("known_storage_gb", 0.0)])
    totals_sheet.append(["Start Date Filter", start_date])
    totals_sheet.append(["End Date Filter", end_date])

    for ws in workbook.worksheets:
        style_header_row(ws)
        autofit_columns(ws)

    workbook.save(xlsx_path)
    return csv_path, xlsx_path, totals
